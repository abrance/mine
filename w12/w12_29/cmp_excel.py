import xlrd
from pathlib import Path
import openpyxl

"""
1.TO表 给 fix 内容：产品名称（去前面的数字+_）6、产品代码（真实代码）8、成立日期9、成立规模（资产净值）10、
"""

f2 = 'fix.xlsx'
f1 = 'TO中金财富产品-资产净值及两费浏览表-20201225.xls'
file = str((Path('.')/f1).resolve())
to_file = str((Path('.')/f2).resolve())
a_file = str((Path('.')/'test.xlsx').resolve())

f1_workbook = xlrd.open_workbook(file)  # 打开工作簿
f1_sheets = f1_workbook.sheet_names()  # 获取工作簿中的所有表格
f1_worksheet = f1_workbook.sheet_by_name(f1_sheets[0])  # 获取工作簿中所有表格中的的第一个表格
f1_row_num = f1_worksheet.nrows
f1_col_num = f1_worksheet.ncols

f2_workbook = xlrd.open_workbook(to_file)  # 打开工作簿
f2_sheets = f2_workbook.sheet_names()  # 获取工作簿中的所有表格
f2_worksheet = f2_workbook.sheet_by_name(f2_sheets[0])  # 获取工作簿中所有表格中的的第一个表格
f2_row_num = f2_worksheet.nrows
f2_col_num = f2_worksheet.ncols


def get_index(sheet, name):
    row_num = sheet.nrows
    col_num = sheet.ncols
    for i in range(0, col_num):

        val = sheet.cell_value(0, i)
        if val == name:
            return i


# 修订表关注 产品名称 产品代码 成立日期 成立规模
pro_name_i = get_index(f2_worksheet, '产品名称')
pro_code_i = get_index(f2_worksheet, '产品代码')
date_i = get_index(f2_worksheet, '成立日期')
scale_i = get_index(f2_worksheet, '成立规模')
#
# # TO表关注 账套名称 产品代码 真实代码 成立日期 资产净值
# name_i1 = get_index(f1_worksheet, '账套名称')
# rel_code_i1 = get_index(f1_worksheet, '真实代码')
# date_i1 = get_index(f1_worksheet, '成立日期')
# val_i1 = get_index(f1_worksheet, '资产净值')

name_i1 = 1
rel_code_i1 = 3
date_i1 = 4
val_i1 = 6


def insert(ls):
    book = openpyxl.load_workbook(to_file)
    names = book.sheetnames
    sheet = book[names[0]]
    for index, name, code, date, val in ls:

        sheet.cell(index, pro_name_i, name)
        sheet.cell(index, pro_code_i, code)
        sheet.cell(index, date_i, date)
        sheet.cell(index, scale_i, val)

    book.save(a_file)


def read():

    append_num = 700

    ls = []
    for i in range(f1_col_num):
        code_f1 = f1_worksheet.cell(i, rel_code_i1).value

        if not code_f1:
            continue

        # find
        for j in range(f2_col_num):
            if code_f1 == f2_worksheet.cell_value(j, pro_code_i):
                break
        else:
            name = f1_worksheet.cell_value(i, name_i1)
            code = f1_worksheet.cell_value(i, rel_code_i1)
            date = f1_worksheet.cell_value(i, date_i1)
            val = f1_worksheet.cell_value(i, val_i1)
            ls.append((append_num, name, code, date, val))
            append_num += 1

    insert(ls)
    print('done')


if __name__ == '__main__':
    # read(file)
    # print(digits)
    # read_excel_xls(str(to_file))
    # append()
    # print(name_i1)
    read()
    # print(f1_worksheet.merged_cells)
